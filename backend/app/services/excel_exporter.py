"""
Сервис экспорта данных в XLSX через openpyxl
"""
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os

from app.core.config import settings


class ExcelExporter:
    """Экспортёр данных в XLSX"""
    
    @staticmethod
    async def export_objects(data: list[dict[str, Any]], filename: str = None) -> str:
        """Экспорт объектов в XLSX"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Объекты"
        
        # Заголовки
        headers = ["ID", "Тип", "Адрес", "Город", "Район", "Статус", "Визитов", "Последний визит"]
        ws.append(headers)
        
        # Стиль заголовков
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Данные
        for row_idx, item in enumerate(data, start=2):
            ws.append([
                str(item.get("id", "")),
                item.get("type", ""),
                item.get("address", ""),
                item.get("city", ""),
                item.get("district", ""),
                item.get("status", ""),
                item.get("visits_count", 0),
                item.get("last_visit_at", ""),
            ])
        
        # Автоширина колонок
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
        
        # Сохраняем файл
        if not filename:
            filename = f"objects_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        filepath = os.path.join(settings.REPORTS_PATH, filename)
        os.makedirs(settings.REPORTS_PATH, exist_ok=True)
        
        wb.save(filepath)
        return filepath

